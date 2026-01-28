"""
Bheem Workspace - Unified Spreadsheet Service
==============================================
Handles spreadsheet operations for both Internal (ERP) and External (SaaS) modes.
Integrates with OnlyOffice Document Server for full Excel-compatible editing.

Storage Architecture:
- Internal mode: S3 at internal/{company_id}/spreadsheets/
- External mode: S3 at external/{tenant_id}/spreadsheets/
- Nextcloud sync: Optional WebDAV sync for user access

Features:
- XLSX file-based storage (not JSON blobs)
- OnlyOffice real-time collaboration
- Version control
- ERP entity linking (internal mode)
- Nextcloud integration
"""

import logging
import hashlib
import uuid
import jwt
import time
from datetime import datetime, timedelta
from enum import Enum
from io import BytesIO
from typing import Optional, Dict, Any, List, Tuple
from uuid import UUID

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from core.config import settings
from services.docs_storage_service import DocsStorageService, get_docs_storage_service
from services.nextcloud_service import NextcloudService

logger = logging.getLogger(__name__)


class SpreadsheetMode(str, Enum):
    """Spreadsheet storage mode"""
    INTERNAL = "internal"  # ERP mode - uses company_id, entity linking
    EXTERNAL = "external"  # SaaS mode - uses tenant_id, simple sharing


class SpreadsheetService:
    """
    Unified spreadsheet service for dual-mode architecture.

    Internal Mode (ERP):
    - Storage: S3 at internal/{company_id}/spreadsheets/
    - Database: workspace.spreadsheets with storage_mode='internal'
    - Features: ERP entity linking, versioning, audit
    - Users: ERP employees synced from HR

    External Mode (SaaS):
    - Storage: S3 at external/{tenant_id}/spreadsheets/
    - Database: workspace.spreadsheets with storage_mode='external'
    - Features: Sharing, collaboration, public links
    - Users: Self-registered SaaS customers
    """

    def __init__(self, db: AsyncSession):
        self.db = db
        self.storage = get_docs_storage_service()
        self.nextcloud = NextcloudService() if settings.PROVISIONING_SYNC_NEXTCLOUD else None

    # =============================================
    # Core Operations
    # =============================================

    async def create_spreadsheet(
        self,
        title: str,
        mode: SpreadsheetMode,
        tenant_id: UUID,
        user_id: UUID,
        company_id: Optional[UUID] = None,
        description: Optional[str] = None,
        folder_id: Optional[UUID] = None,
        template_id: Optional[str] = None,
        linked_entity_type: Optional[str] = None,
        linked_entity_id: Optional[UUID] = None,
    ) -> Dict[str, Any]:
        """
        Create a new spreadsheet as XLSX file.

        Args:
            title: Spreadsheet title
            mode: Internal (ERP) or External (SaaS)
            tenant_id: Workspace tenant ID
            user_id: Creator's user ID
            company_id: ERP company ID (required for internal mode)
            description: Optional description
            folder_id: Optional folder for organization
            template_id: Optional template to use
            linked_entity_type: ERP entity type (internal mode)
            linked_entity_id: ERP entity ID (internal mode)

        Returns:
            Spreadsheet metadata with edit URL
        """
        spreadsheet_id = uuid.uuid4()
        document_key = self._generate_document_key(spreadsheet_id)

        # Generate XLSX content
        if template_id:
            xlsx_bytes = await self._get_template_content(template_id)
        else:
            xlsx_bytes = self._create_empty_xlsx(title)

        # Calculate checksum
        checksum = hashlib.sha256(xlsx_bytes).hexdigest()
        file_size = len(xlsx_bytes)

        # Determine storage path based on mode
        if mode == SpreadsheetMode.INTERNAL:
            if not company_id:
                raise ValueError("company_id is required for internal mode")
            storage_path = f"internal/{company_id}/spreadsheets/{spreadsheet_id}.xlsx"
        else:
            storage_path = f"external/{tenant_id}/spreadsheets/{spreadsheet_id}.xlsx"

        # Upload to S3
        try:
            await self.storage.upload_file(
                file=BytesIO(xlsx_bytes),
                filename=f"{spreadsheet_id}.xlsx",
                company_id=company_id if mode == SpreadsheetMode.INTERNAL else None,
                tenant_id=tenant_id if mode == SpreadsheetMode.EXTERNAL else None,
                folder_path="spreadsheets",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            logger.info(f"Uploaded spreadsheet to S3: {storage_path}")
        except Exception as e:
            logger.error(f"Failed to upload spreadsheet to S3: {e}")
            raise

        # Optionally sync to Nextcloud
        nextcloud_path = None
        if self.nextcloud:
            try:
                nextcloud_path = f"/Spreadsheets/{title}.xlsx"
                # Note: Would need user credentials for Nextcloud upload
                # await self._sync_to_nextcloud(user_id, nextcloud_path, xlsx_bytes)
            except Exception as e:
                logger.warning(f"Nextcloud sync failed: {e}")

        # Insert into database
        query = text("""
            INSERT INTO workspace.spreadsheets (
                id, tenant_id, title, description, folder_id,
                storage_path, storage_bucket, file_size, checksum,
                nextcloud_path, version, storage_mode, document_key,
                linked_entity_type, linked_entity_id,
                created_by, created_at, updated_at
            ) VALUES (
                :id, :tenant_id, :title, :description, :folder_id,
                :storage_path, :storage_bucket, :file_size, :checksum,
                :nextcloud_path, 1, :storage_mode, :document_key,
                :linked_entity_type, :linked_entity_id,
                :user_id, NOW(), NOW()
            )
            RETURNING id, title, storage_path, version, document_key, created_at
        """)

        result = await self.db.execute(query, {
            "id": spreadsheet_id,
            "tenant_id": tenant_id,
            "title": title,
            "description": description,
            "folder_id": folder_id,
            "storage_path": storage_path,
            "storage_bucket": self.storage.bucket,
            "file_size": file_size,
            "checksum": checksum,
            "nextcloud_path": nextcloud_path,
            "storage_mode": mode.value,
            "document_key": document_key,
            "linked_entity_type": linked_entity_type,
            "linked_entity_id": linked_entity_id,
            "user_id": user_id,
        })
        await self.db.commit()

        row = result.fetchone()

        # Create initial version record
        await self._create_version_record(
            spreadsheet_id=spreadsheet_id,
            version_number=1,
            storage_path=storage_path,
            file_size=file_size,
            checksum=checksum,
            user_id=user_id,
            comment="Initial version"
        )

        return {
            "id": str(spreadsheet_id),
            "title": title,
            "description": description,
            "storage_path": storage_path,
            "storage_mode": mode.value,
            "version": 1,
            "document_key": document_key,
            "file_size": file_size,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "edit_url": f"/sheets/{spreadsheet_id}/edit",
        }

    async def get_spreadsheet(
        self,
        spreadsheet_id: UUID,
        tenant_id: UUID,
        user_id: UUID,
    ) -> Optional[Dict[str, Any]]:
        """Get spreadsheet metadata."""
        query = text("""
            SELECT
                s.id, s.title, s.description, s.folder_id,
                s.storage_path, s.storage_bucket, s.file_size, s.checksum,
                s.nextcloud_path, s.version, s.storage_mode, s.document_key,
                s.linked_entity_type, s.linked_entity_id,
                s.is_starred, s.is_deleted,
                s.created_by, s.created_at, s.updated_at,
                s.last_edited_by, s.last_edited_at,
                tu.email as creator_email,
                tu.name as creator_name
            FROM workspace.spreadsheets s
            LEFT JOIN workspace.tenant_users tu ON s.created_by = tu.id
            WHERE s.id = :spreadsheet_id
            AND s.tenant_id = :tenant_id
            AND s.is_deleted = FALSE
        """)

        result = await self.db.execute(query, {
            "spreadsheet_id": spreadsheet_id,
            "tenant_id": tenant_id,
        })
        row = result.fetchone()

        if not row:
            return None

        return {
            "id": str(row.id),
            "title": row.title,
            "description": row.description,
            "folder_id": str(row.folder_id) if row.folder_id else None,
            "storage_path": row.storage_path,
            "storage_bucket": row.storage_bucket,
            "file_size": row.file_size,
            "checksum": row.checksum,
            "nextcloud_path": row.nextcloud_path,
            "version": row.version,
            "storage_mode": row.storage_mode,
            "document_key": row.document_key,
            "linked_entity_type": row.linked_entity_type,
            "linked_entity_id": str(row.linked_entity_id) if row.linked_entity_id else None,
            "is_starred": row.is_starred,
            "created_by": str(row.created_by),
            "creator_email": row.creator_email,
            "creator_name": row.creator_name,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
            "last_edited_by": str(row.last_edited_by) if row.last_edited_by else None,
            "last_edited_at": row.last_edited_at.isoformat() if row.last_edited_at else None,
        }

    async def list_spreadsheets(
        self,
        tenant_id: UUID,
        user_id: UUID,
        folder_id: Optional[UUID] = None,
        search: Optional[str] = None,
        starred_only: bool = False,
        include_deleted: bool = False,
        limit: int = 50,
        offset: int = 0,
    ) -> Dict[str, Any]:
        """List spreadsheets with filtering."""
        conditions = ["s.tenant_id = :tenant_id"]
        params = {"tenant_id": tenant_id, "limit": limit, "offset": offset}

        if not include_deleted:
            conditions.append("s.is_deleted = FALSE")

        if folder_id:
            conditions.append("s.folder_id = :folder_id")
            params["folder_id"] = folder_id

        if search:
            conditions.append("s.title ILIKE :search")
            params["search"] = f"%{search}%"

        if starred_only:
            conditions.append("s.is_starred = TRUE")

        where_clause = " AND ".join(conditions)

        query = text(f"""
            SELECT
                s.id, s.title, s.description, s.folder_id,
                s.storage_mode, s.version, s.file_size,
                s.is_starred, s.created_by, s.created_at, s.updated_at,
                tu.email as creator_email,
                tu.name as creator_name
            FROM workspace.spreadsheets s
            LEFT JOIN workspace.tenant_users tu ON s.created_by = tu.id
            WHERE {where_clause}
            ORDER BY s.updated_at DESC
            LIMIT :limit OFFSET :offset
        """)

        result = await self.db.execute(query, params)
        rows = result.fetchall()

        # Get total count
        count_query = text(f"""
            SELECT COUNT(*) FROM workspace.spreadsheets s
            WHERE {where_clause}
        """)
        count_result = await self.db.execute(count_query, params)
        total = count_result.scalar()

        spreadsheets = []
        for row in rows:
            spreadsheets.append({
                "id": str(row.id),
                "title": row.title,
                "description": row.description,
                "folder_id": str(row.folder_id) if row.folder_id else None,
                "storage_mode": row.storage_mode,
                "version": row.version,
                "file_size": row.file_size,
                "is_starred": row.is_starred,
                "created_by": str(row.created_by),
                "creator_email": row.creator_email,
                "creator_name": row.creator_name,
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "updated_at": row.updated_at.isoformat() if row.updated_at else None,
            })

        return {
            "spreadsheets": spreadsheets,
            "total": total,
            "limit": limit,
            "offset": offset,
            "has_more": offset + len(spreadsheets) < total,
        }

    # =============================================
    # OnlyOffice Integration
    # =============================================

    async def get_editor_config(
        self,
        spreadsheet_id: UUID,
        tenant_id: UUID,
        user_id: UUID,
        user_name: str,
        user_email: str,
        mode: str = "edit",  # edit, view, review
    ) -> Dict[str, Any]:
        """
        Get OnlyOffice Document Editor configuration.

        This generates the config object that the OnlyOffice JS API needs
        to initialize the editor.

        Args:
            spreadsheet_id: Spreadsheet UUID
            tenant_id: Tenant UUID for access validation
            user_id: User UUID
            user_name: Display name for collaboration
            user_email: User email
            mode: Edit mode (edit, view, review)

        Returns:
            OnlyOffice editor configuration
        """
        # Get spreadsheet metadata
        spreadsheet = await self.get_spreadsheet(spreadsheet_id, tenant_id, user_id)
        if not spreadsheet:
            raise ValueError("Spreadsheet not found")

        # Check if storage_path exists, if not create the XLSX file
        storage_path = spreadsheet.get("storage_path")
        if not storage_path:
            # Create empty XLSX file for legacy spreadsheets
            storage_path = await self._create_xlsx_for_legacy_spreadsheet(
                spreadsheet_id=spreadsheet_id,
                tenant_id=tenant_id,
                user_id=user_id,
                title=spreadsheet.get("title", "Untitled"),
            )
            spreadsheet["storage_path"] = storage_path

        # Generate document key (unique per version for cache invalidation)
        document_key = f"{spreadsheet_id}-v{spreadsheet['version']}-{int(time.time())}"

        # Generate access token for document proxy endpoint
        access_token = jwt.encode(
            {
                "sheet_id": str(spreadsheet_id),
                "exp": datetime.utcnow() + timedelta(hours=24),
            },
            settings.ONLYOFFICE_JWT_SECRET,
            algorithm="HS256"
        )

        # Ensure token is a string (PyJWT 2.0+ returns string, older versions returned bytes)
        if isinstance(access_token, bytes):
            access_token = access_token.decode('utf-8')

        # URL-encode the token for safety (even though base64url should be URL-safe)
        from urllib.parse import quote
        encoded_token = quote(access_token, safe='')

        # Use proxy URL instead of S3 presigned URL (OnlyOffice can't access S3 directly)
        # The proxy endpoint streams the file from S3 through our backend
        download_url = f"{settings.WORKSPACE_URL}/api/v1/sheets/{spreadsheet_id}/content?token={encoded_token}"

        logger.info(f"Generated download URL for spreadsheet {spreadsheet_id}: {download_url[:100]}...")

        # Update document key in database
        await self.db.execute(
            text("UPDATE workspace.spreadsheets SET document_key = :key WHERE id = :id"),
            {"key": document_key, "id": spreadsheet_id}
        )
        await self.db.commit()

        # Callback URL for OnlyOffice to notify us of changes
        callback_url = f"{settings.ONLYOFFICE_CALLBACK_URL}/{spreadsheet_id}/onlyoffice-callback"

        # Build editor config
        config = {
            "document": {
                "fileType": "xlsx",
                "key": document_key,
                "title": f"{spreadsheet['title']}.xlsx",
                "url": download_url,
                "permissions": {
                    "chat": True,  # Chat permission (moved from customization)
                    "comment": True,
                    "copy": True,
                    "download": True,
                    "edit": mode == "edit",
                    "print": True,
                    "review": mode in ["edit", "review"],
                },
            },
            "documentType": "cell",  # spreadsheet type
            "editorConfig": {
                "callbackUrl": callback_url,
                "lang": "en",
                "mode": mode,
                "user": {
                    "id": str(user_id),
                    "name": user_name,
                },
                "customization": {
                    "autosave": True,
                    "comments": True,
                    "compactHeader": False,
                    "compactToolbar": False,
                    "feedback": False,
                    "forcesave": True,
                    "help": False,
                    "hideRightMenu": False,
                    "toolbarNoTabs": False,
                    "toolbarHideFileName": False,
                    "statusBar": True,
                    "leftMenu": True,
                    "rightMenu": True,
                    "toolbar": True,
                    "zoom": 100,
                    "logo": {
                        "image": f"{settings.WORKSPACE_URL}/static/bheem-logo.svg",
                        "imageDark": f"{settings.WORKSPACE_URL}/static/bheem-logo-dark.svg",
                        "imageLight": f"{settings.WORKSPACE_URL}/static/bheem-logo-light.svg",
                        "imageEmbedded": f"{settings.WORKSPACE_URL}/static/bheem-loader-logo.svg",
                        "url": f"{settings.WORKSPACE_URL}/sheets",
                        "visible": True
                    },
                    "customer": {
                        "address": "Bheem Cloud Services",
                        "info": "Bheem Workspace - Your productivity suite",
                        "logo": f"{settings.WORKSPACE_URL}/static/bheem-logo.svg",
                        "logoDark": f"{settings.WORKSPACE_URL}/static/bheem-logo-dark.svg",
                        "mail": "support@bheem.cloud",
                        "name": "Bheem Sheets",
                        "www": "https://bheem.cloud"
                    },
                    "goback": {
                        "blank": False,
                        "text": "Back to Bheem",
                        "url": f"{settings.WORKSPACE_URL}/sheets"
                    },
                    "loaderLogo": f"{settings.WORKSPACE_URL}/static/bheem-loader-logo.svg",
                    "loaderName": "Bheem Sheets",
                },
            },
            "height": "100%",
            "width": "100%",
            "type": "desktop",
        }

        # Add JWT token if enabled
        if settings.ONLYOFFICE_JWT_ENABLED and settings.ONLYOFFICE_JWT_SECRET:
            token = self._generate_onlyoffice_jwt(config)
            config["token"] = token

        # Record edit session
        await self._record_edit_session(
            spreadsheet_id=spreadsheet_id,
            user_id=user_id,
            document_key=document_key,
        )

        return {
            "config": config,
            "documentServerUrl": settings.ONLYOFFICE_URL,
            "spreadsheet": spreadsheet,
        }

    async def handle_onlyoffice_callback(
        self,
        spreadsheet_id: UUID,
        callback_data: Dict[str, Any],
    ) -> Dict[str, int]:
        """
        Handle OnlyOffice Document Server callback.

        OnlyOffice calls this endpoint when document state changes.

        Status codes:
        - 0: No document change
        - 1: Document being edited
        - 2: Document ready for saving (users closed editors)
        - 3: Document saving error
        - 4: Document closed with no changes
        - 6: Document being edited, force save requested
        - 7: Error force saving document

        Args:
            spreadsheet_id: Spreadsheet UUID
            callback_data: OnlyOffice callback payload

        Returns:
            {"error": 0} for success, {"error": 1} for failure
        """
        status = callback_data.get("status")
        document_key = callback_data.get("key")
        users = callback_data.get("users", [])

        logger.info(f"OnlyOffice callback for {spreadsheet_id}: status={status}, key={document_key}")

        try:
            if status == 2 or status == 6:
                # Document ready for saving or force save
                document_url = callback_data.get("url")

                if not document_url:
                    logger.error("No document URL in callback")
                    return {"error": 1}

                # Download the edited document from OnlyOffice
                async with httpx.AsyncClient(verify=False) as client:
                    response = await client.get(document_url, timeout=60.0)
                    if response.status_code != 200:
                        logger.error(f"Failed to download from OnlyOffice: {response.status_code}")
                        return {"error": 1}
                    new_xlsx_bytes = response.content

                # Get current spreadsheet info
                result = await self.db.execute(
                    text("""
                        SELECT storage_path, storage_bucket, version, tenant_id, created_by
                        FROM workspace.spreadsheets
                        WHERE id = :id
                    """),
                    {"id": spreadsheet_id}
                )
                row = result.fetchone()

                if not row:
                    logger.error(f"Spreadsheet not found: {spreadsheet_id}")
                    return {"error": 1}

                storage_path = row.storage_path
                current_version = row.version
                tenant_id = row.tenant_id

                # Calculate new checksum
                new_checksum = hashlib.sha256(new_xlsx_bytes).hexdigest()
                new_file_size = len(new_xlsx_bytes)

                # Upload new version to S3 (overwrite)
                await self.storage.upload_file(
                    file=BytesIO(new_xlsx_bytes),
                    filename=storage_path.split("/")[-1],
                    company_id=None,
                    tenant_id=tenant_id,  # Use tenant_id from spreadsheet record
                    folder_path="spreadsheets",
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                # Increment version
                new_version = current_version + 1

                # Update database
                await self.db.execute(
                    text("""
                        UPDATE workspace.spreadsheets
                        SET version = :version,
                            file_size = :file_size,
                            checksum = :checksum,
                            updated_at = NOW(),
                            last_edited_at = NOW()
                        WHERE id = :id
                    """),
                    {
                        "id": spreadsheet_id,
                        "version": new_version,
                        "file_size": new_file_size,
                        "checksum": new_checksum,
                    }
                )

                # Create version record
                await self._create_version_record(
                    spreadsheet_id=spreadsheet_id,
                    version_number=new_version,
                    storage_path=storage_path,
                    file_size=new_file_size,
                    checksum=new_checksum,
                    user_id=row.created_by,  # Should be last editor
                    comment=f"Auto-saved version {new_version}"
                )

                await self.db.commit()

                logger.info(f"Saved spreadsheet {spreadsheet_id} version {new_version}")

            elif status == 4:
                # Document closed with no changes
                logger.info(f"Spreadsheet {spreadsheet_id} closed without changes")

            elif status == 1:
                # Document being edited - update session
                await self._update_edit_session(spreadsheet_id, document_key)

            return {"error": 0}

        except Exception as e:
            logger.error(f"OnlyOffice callback error: {e}")
            await self.db.rollback()
            return {"error": 1}

    # =============================================
    # Version Control
    # =============================================

    async def get_versions(
        self,
        spreadsheet_id: UUID,
        tenant_id: UUID,
        limit: int = 20,
    ) -> List[Dict[str, Any]]:
        """Get version history for a spreadsheet."""
        query = text("""
            SELECT
                v.id, v.version_number, v.storage_path, v.file_size, v.checksum,
                v.created_by, v.created_at, v.comment, v.is_current,
                tu.name as creator_name, tu.email as creator_email
            FROM workspace.spreadsheet_versions v
            LEFT JOIN workspace.tenant_users tu ON v.created_by = tu.id
            JOIN workspace.spreadsheets s ON v.spreadsheet_id = s.id
            WHERE v.spreadsheet_id = :spreadsheet_id
            AND s.tenant_id = :tenant_id
            ORDER BY v.version_number DESC
            LIMIT :limit
        """)

        result = await self.db.execute(query, {
            "spreadsheet_id": spreadsheet_id,
            "tenant_id": tenant_id,
            "limit": limit,
        })

        versions = []
        for row in result.fetchall():
            versions.append({
                "id": str(row.id),
                "version_number": row.version_number,
                "file_size": row.file_size,
                "checksum": row.checksum,
                "created_by": str(row.created_by) if row.created_by else None,
                "creator_name": row.creator_name,
                "creator_email": row.creator_email,
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "comment": row.comment,
                "is_current": row.is_current,
            })

        return versions

    async def restore_version(
        self,
        spreadsheet_id: UUID,
        version_number: int,
        tenant_id: UUID,
        user_id: UUID,
    ) -> Dict[str, Any]:
        """Restore a previous version of a spreadsheet."""
        # Get version to restore
        result = await self.db.execute(
            text("""
                SELECT v.storage_path, v.file_size, v.checksum
                FROM workspace.spreadsheet_versions v
                JOIN workspace.spreadsheets s ON v.spreadsheet_id = s.id
                WHERE v.spreadsheet_id = :spreadsheet_id
                AND v.version_number = :version_number
                AND s.tenant_id = :tenant_id
            """),
            {
                "spreadsheet_id": spreadsheet_id,
                "version_number": version_number,
                "tenant_id": tenant_id,
            }
        )
        version_row = result.fetchone()

        if not version_row:
            raise ValueError("Version not found")

        # Get current spreadsheet
        result = await self.db.execute(
            text("SELECT version, storage_path FROM workspace.spreadsheets WHERE id = :id"),
            {"id": spreadsheet_id}
        )
        current = result.fetchone()

        # Create new version number
        new_version = current.version + 1

        # Update spreadsheet to point to restored version's data
        await self.db.execute(
            text("""
                UPDATE workspace.spreadsheets
                SET version = :new_version,
                    file_size = :file_size,
                    checksum = :checksum,
                    updated_at = NOW(),
                    last_edited_by = :user_id,
                    last_edited_at = NOW()
                WHERE id = :id
            """),
            {
                "id": spreadsheet_id,
                "new_version": new_version,
                "file_size": version_row.file_size,
                "checksum": version_row.checksum,
                "user_id": user_id,
            }
        )

        # Create version record
        await self._create_version_record(
            spreadsheet_id=spreadsheet_id,
            version_number=new_version,
            storage_path=version_row.storage_path,
            file_size=version_row.file_size,
            checksum=version_row.checksum,
            user_id=user_id,
            comment=f"Restored from version {version_number}"
        )

        await self.db.commit()

        return {
            "spreadsheet_id": str(spreadsheet_id),
            "restored_from": version_number,
            "new_version": new_version,
        }

    # =============================================
    # ERP Entity Linking (Internal Mode)
    # =============================================

    async def link_to_entity(
        self,
        spreadsheet_id: UUID,
        entity_type: str,
        entity_id: UUID,
        tenant_id: UUID,
        user_id: UUID,
    ) -> Dict[str, Any]:
        """
        Link spreadsheet to ERP entity (internal mode only).

        Supported entity types:
        - SALES_INVOICE, SALES_ORDER, SALES_QUOTE
        - PURCHASE_ORDER, PURCHASE_BILL
        - CUSTOMER, VENDOR
        - PROJECT, TASK
        - EMPLOYEE
        """
        # Verify spreadsheet exists and is internal mode
        result = await self.db.execute(
            text("""
                SELECT storage_mode FROM workspace.spreadsheets
                WHERE id = :id AND tenant_id = :tenant_id
            """),
            {"id": spreadsheet_id, "tenant_id": tenant_id}
        )
        row = result.fetchone()

        if not row:
            raise ValueError("Spreadsheet not found")

        if row.storage_mode != "internal":
            raise ValueError("Entity linking only available for internal mode spreadsheets")

        # Update spreadsheet with entity link
        await self.db.execute(
            text("""
                UPDATE workspace.spreadsheets
                SET linked_entity_type = :entity_type,
                    linked_entity_id = :entity_id,
                    updated_at = NOW()
                WHERE id = :id
            """),
            {
                "id": spreadsheet_id,
                "entity_type": entity_type,
                "entity_id": entity_id,
            }
        )
        await self.db.commit()

        return {
            "spreadsheet_id": str(spreadsheet_id),
            "linked_entity_type": entity_type,
            "linked_entity_id": str(entity_id),
        }

    # =============================================
    # Helper Methods
    # =============================================

    def _create_empty_xlsx(self, title: str) -> bytes:
        """Create empty XLSX workbook."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Set default column widths (A-Z)
        for col in range(1, 27):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Set default row heights (1-100)
        for row in range(1, 101):
            ws.row_dimensions[row].height = 20

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    async def _get_template_content(self, template_id: str) -> bytes:
        """Get template content from database or storage."""
        result = await self.db.execute(
            text("""
                SELECT content FROM workspace.productivity_templates
                WHERE id = :id AND template_type = 'spreadsheet'
            """),
            {"id": template_id}
        )
        row = result.fetchone()

        if row and row.content:
            # If template has storage path, download from S3
            if isinstance(row.content, dict) and "storage_path" in row.content:
                file_content, _ = await self.storage.download_file(row.content["storage_path"])
                return file_content.read()

        # Fall back to empty workbook
        return self._create_empty_xlsx("Untitled spreadsheet")

    def _generate_document_key(self, spreadsheet_id: UUID) -> str:
        """Generate unique document key for OnlyOffice."""
        timestamp = int(time.time() * 1000)
        return f"{spreadsheet_id}-{timestamp}"

    def _generate_onlyoffice_jwt(self, config: Dict[str, Any]) -> str:
        """Generate JWT token for OnlyOffice."""
        payload = config.copy()
        payload["exp"] = datetime.utcnow() + timedelta(hours=24)

        return jwt.encode(
            payload,
            settings.ONLYOFFICE_JWT_SECRET,
            algorithm="HS256"
        )

    async def _create_xlsx_for_legacy_spreadsheet(
        self,
        spreadsheet_id: UUID,
        tenant_id: UUID,
        user_id: UUID,
        title: str,
    ) -> str:
        """
        Create XLSX file for legacy spreadsheets that don't have storage_path.

        This handles spreadsheets created before the OnlyOffice integration
        that only have JSON data in worksheets table.
        """
        from openpyxl import Workbook
        import hashlib

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Try to load existing worksheet data (JSON format)
        try:
            result = await self.db.execute(
                text("""
                    SELECT name, data, sheet_index
                    FROM workspace.worksheets
                    WHERE spreadsheet_id = :spreadsheet_id
                    ORDER BY sheet_index
                """),
                {"spreadsheet_id": spreadsheet_id}
            )
            worksheets = result.fetchall()

            if worksheets:
                # Convert JSON data to Excel
                for i, worksheet in enumerate(worksheets):
                    if i == 0:
                        ws.title = worksheet.name or "Sheet1"
                    else:
                        ws = wb.create_sheet(title=worksheet.name or f"Sheet{i+1}")

                    # Parse JSON data if exists
                    if worksheet.data:
                        data = worksheet.data if isinstance(worksheet.data, dict) else {}
                        for cell_ref, cell_data in data.items():
                            try:
                                if isinstance(cell_data, dict):
                                    ws[cell_ref] = cell_data.get("value", "")
                                else:
                                    ws[cell_ref] = cell_data
                            except Exception:
                                pass
        except Exception as e:
            logger.warning(f"Could not load worksheet data: {e}")

        # Save to bytes
        xlsx_buffer = BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)  # Reset to beginning for reading
        xlsx_bytes = xlsx_buffer.getvalue()

        # Calculate checksum and size
        file_size = len(xlsx_bytes)
        checksum = hashlib.sha256(xlsx_bytes).hexdigest()

        # Generate filename and storage path
        filename = f"{spreadsheet_id}.xlsx"

        # Upload to S3 using the storage service
        xlsx_buffer.seek(0)  # Reset again for upload
        upload_result = await self.storage.upload_file(
            file=xlsx_buffer,
            filename=filename,
            tenant_id=tenant_id,
            folder_path="spreadsheets",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Get storage path from upload result
        storage_path = upload_result.get("storage_path", f"spreadsheets/{tenant_id}/{filename}")

        # Update spreadsheet record
        await self.db.execute(
            text("""
                UPDATE workspace.spreadsheets
                SET storage_path = :storage_path,
                    storage_bucket = :bucket,
                    file_size = :file_size,
                    checksum = :checksum,
                    version = COALESCE(version, 1),
                    updated_at = NOW()
                WHERE id = :spreadsheet_id
            """),
            {
                "spreadsheet_id": spreadsheet_id,
                "storage_path": storage_path,
                "bucket": self.storage.bucket,
                "file_size": file_size,
                "checksum": checksum,
            }
        )
        await self.db.commit()

        logger.info(f"Created XLSX file for legacy spreadsheet {spreadsheet_id}")
        return storage_path

    async def _create_version_record(
        self,
        spreadsheet_id: UUID,
        version_number: int,
        storage_path: str,
        file_size: int,
        checksum: str,
        user_id: UUID,
        comment: Optional[str] = None,
    ):
        """Create version history record."""
        # Mark previous versions as not current
        await self.db.execute(
            text("""
                UPDATE workspace.spreadsheet_versions
                SET is_current = FALSE
                WHERE spreadsheet_id = :spreadsheet_id
            """),
            {"spreadsheet_id": spreadsheet_id}
        )

        # Insert new version
        await self.db.execute(
            text("""
                INSERT INTO workspace.spreadsheet_versions (
                    id, spreadsheet_id, version_number, storage_path,
                    file_size, checksum, created_by, created_at, comment, is_current
                ) VALUES (
                    :id, :spreadsheet_id, :version_number, :storage_path,
                    :file_size, :checksum, :user_id, NOW(), :comment, TRUE
                )
            """),
            {
                "id": uuid.uuid4(),
                "spreadsheet_id": spreadsheet_id,
                "version_number": version_number,
                "storage_path": storage_path,
                "file_size": file_size,
                "checksum": checksum,
                "user_id": user_id,
                "comment": comment,
            }
        )

    async def _record_edit_session(
        self,
        spreadsheet_id: UUID,
        user_id: UUID,
        document_key: str,
    ):
        """Record user edit session for tracking."""
        await self.db.execute(
            text("""
                INSERT INTO workspace.spreadsheet_edit_sessions (
                    id, spreadsheet_id, user_id, document_key, started_at, last_activity_at, status
                ) VALUES (
                    :id, :spreadsheet_id, :user_id, :document_key, NOW(), NOW(), 'active'
                )
                ON CONFLICT (document_key, user_id) DO UPDATE
                SET last_activity_at = NOW(), status = 'active'
            """),
            {
                "id": uuid.uuid4(),
                "spreadsheet_id": spreadsheet_id,
                "user_id": user_id,
                "document_key": document_key,
            }
        )
        await self.db.commit()

    async def _update_edit_session(self, spreadsheet_id: UUID, document_key: str):
        """Update edit session activity timestamp."""
        await self.db.execute(
            text("""
                UPDATE workspace.spreadsheet_edit_sessions
                SET last_activity_at = NOW()
                WHERE document_key = :document_key
            """),
            {"document_key": document_key}
        )
        await self.db.commit()


# Singleton instance
_spreadsheet_service: Optional[SpreadsheetService] = None


def get_spreadsheet_service(db: AsyncSession) -> SpreadsheetService:
    """Get spreadsheet service instance."""
    return SpreadsheetService(db)
